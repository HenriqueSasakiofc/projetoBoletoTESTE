import hashlib
import json
import re
import unicodedata
import zipfile
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO
from uuid import uuid4

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from ..config import settings
from .. import models

EMAIL_RE = re.compile(r"^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$")
FORMULA_PREFIXES = ("=", "+", "-", "@")
CUSTOMER_REQUIRED_HINTS = {"NOME", "CODIGO"}
RECEIVABLE_REQUIRED_HINTS = {"NOME", "DOCUMENTO", "VENCIMENTO"}


def normalize_text(value: str | None) -> str:
    if not value:
        return ""
    value = str(value).strip()
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = re.sub(r"\s+", " ", value)
    return value.upper().strip()


def normalize_digits(value: str | None) -> str | None:
    if value is None:
        return None
    digits = re.sub(r"\D", "", str(value))
    return digits or None


def sanitize_cell_value(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float, Decimal)):
        return value
    text = str(value).strip()
    if not text:
        return None
    left = text.lstrip()
    if left.startswith(FORMULA_PREFIXES):
        text = "'" + text
    return re.sub(r"\s+", " ", text).strip()


def parse_email(value: str | None) -> str | None:
    if not value:
        return None
    candidate = str(value).replace("mailto:", " ").replace(";", " ").replace(",", " ").strip().split()
    for item in candidate:
        item = item.strip().lower()
        if EMAIL_RE.match(item):
            return item
    return None


def sha256_bytes(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def validate_filename(filename: str | None) -> str:
    safe_name = (filename or "arquivo.xlsx").strip().replace("\\", "/").split("/")[-1]
    if not safe_name.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Somente arquivos .xlsx são aceitos")
    if safe_name.lower().endswith(".xlsm"):
        raise HTTPException(status_code=400, detail="Arquivos com macro não são aceitos")
    return safe_name


def validate_upload_content(filename: str, content: bytes) -> None:
    validate_filename(filename)
    if not content:
        raise HTTPException(status_code=400, detail="Arquivo vazio")
    if len(content) > settings.upload_max_bytes:
        raise HTTPException(status_code=400, detail="Arquivo excede o tamanho máximo permitido")
    try:
        with zipfile.ZipFile(BytesIO(content)) as archive:
            names = {name.lower() for name in archive.namelist()}
            if any(name.endswith("vbaproject.bin") for name in names):
                raise HTTPException(status_code=400, detail="Arquivo com macro não é permitido")
            if not any(name.endswith("workbook.xml") for name in names):
                raise HTTPException(status_code=400, detail="Arquivo XLSX inválido")
    except zipfile.BadZipFile as exc:
        raise HTTPException(status_code=400, detail="Arquivo XLSX inválido") from exc


def workbook_rows(content: bytes):
    workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    return list(sheet.iter_rows(values_only=True))


def detect_header_row(rows: list[tuple], required_hints: set[str]) -> tuple[int, list[str]]:
    for index, row in enumerate(rows[:8]):
        headers = [normalize_text(value) for value in row]
        if required_hints.issubset(set(headers)):
            return index, [str(value).strip() if value is not None else "" for value in row]
    raise HTTPException(status_code=400, detail="Cabeçalho da planilha não reconhecido")


def rows_to_dicts(rows: list[tuple], required_hints: set[str]) -> list[tuple[int, dict[str, object]]]:
    header_index, header = detect_header_row(rows, required_hints)
    output: list[tuple[int, dict[str, object]]] = []
    for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        data = {}
        if not any(value not in (None, "") for value in row):
            continue
        for col_index, col_name in enumerate(header):
            if not col_name:
                continue
            data[col_name] = sanitize_cell_value(row[col_index] if col_index < len(row) else None)
        output.append((row_number, data))
    return output


def parse_decimal(value) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        if isinstance(value, Decimal):
            return value.quantize(Decimal("0.01"))
        text = str(value).replace("R$", "").replace(".", "").replace(",", ".").strip()
        return Decimal(text).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def parse_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def safe_json(data: dict) -> str:
    def default(obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        if isinstance(obj, Decimal):
            return str(obj)
        return str(obj)

    return json.dumps(data, ensure_ascii=False, default=default)


def normalize_receivable_status(raw: str | None, due_date: date | None, balance_amount: Decimal | None) -> str:
    normalized = normalize_text(raw)
    if normalized in {"PAGO", "PAGO PARCIAL", "QUITADO"} or (balance_amount is not None and balance_amount <= 0):
        return "pago"
    if normalized in {"CANCELADO", "BAIXADO"}:
        return "cancelado"
    if not due_date:
        return "em_aberto"
    today = date.today()
    if due_date < today:
        return "inadimplente"
    if (due_date - today).days <= 7:
        return "vencendo"
    return "em_aberto"


def _customer_validation_errors(payload: dict) -> list[str]:
    errors: list[str] = []
    if not payload.get("Nome"):
        errors.append("Nome é obrigatório")
    if payload.get("E-mail para cobrança") and not parse_email(payload.get("E-mail para cobrança")):
        errors.append("E-mail para cobrança inválido")
    if payload.get("E-mail do financeiro") and not parse_email(payload.get("E-mail do financeiro")):
        errors.append("E-mail do financeiro inválido")
    return errors


def _receivable_validation_errors(payload: dict) -> list[str]:
    errors: list[str] = []
    if not payload.get("Nome"):
        errors.append("Nome do cliente é obrigatório")
    if not payload.get("Documento") and not payload.get("Nosso Numero"):
        errors.append("Documento ou Nosso Numero é obrigatório")
    if not parse_date(payload.get("Vencimento")):
        errors.append("Vencimento inválido")
    if parse_decimal(payload.get("Valor")) is None:
        errors.append("Valor inválido")
    return errors


def _clear_batch_staging(db: Session, batch_id: int) -> None:
    db.query(models.CustomerLinkPending).filter(models.CustomerLinkPending.upload_batch_id == batch_id).delete()
    db.query(models.StagingReceivable).filter(models.StagingReceivable.upload_batch_id == batch_id).delete()
    db.query(models.StagingCustomer).filter(models.StagingCustomer.upload_batch_id == batch_id).delete()


def create_upload_batch(
    db: Session,
    *,
    user: models.User,
    clients_filename: str,
    clients_bytes: bytes,
    receivables_filename: str,
    receivables_bytes: bytes,
) -> models.UploadBatch:
    validate_upload_content(clients_filename, clients_bytes)
    validate_upload_content(receivables_filename, receivables_bytes)

    batch = models.UploadBatch(
        company_id=user.company_id,
        uploaded_by_user_id=user.id,
        batch_reference=uuid4().hex,
        clients_filename=validate_filename(clients_filename),
        receivables_filename=validate_filename(receivables_filename),
        clients_file_hash=sha256_bytes(clients_bytes),
        receivables_file_hash=sha256_bytes(receivables_bytes),
        status="uploaded",
    )
    db.add(batch)
    db.flush()

    stage_batch_files(db, batch, clients_bytes, receivables_bytes)
    db.commit()
    db.refresh(batch)
    return batch


def stage_batch_files(db: Session, batch: models.UploadBatch, clients_bytes: bytes, receivables_bytes: bytes) -> None:
    _clear_batch_staging(db, batch.id)

    customer_rows = rows_to_dicts(workbook_rows(clients_bytes), CUSTOMER_REQUIRED_HINTS)
    receivable_rows = rows_to_dicts(workbook_rows(receivables_bytes), RECEIVABLE_REQUIRED_HINTS)

    preview_errors: list[str] = []
    valid_customer_rows = 0
    valid_receivable_rows = 0

    customer_docs: set[str] = set()
    customer_codes: set[str] = set()

    for row_number, payload in customer_rows:
        errors = _customer_validation_errors(payload)
        document = normalize_digits(payload.get("CNPJ/CPF"))
        external_code = str(payload.get("Código") or "").strip() or None
        if document and document in customer_docs:
            errors.append("Documento duplicado na planilha de clientes")
        if external_code and external_code in customer_codes:
            errors.append("Código duplicado na planilha de clientes")
        if document:
            customer_docs.add(document)
        if external_code:
            customer_codes.add(external_code)
        validation_status = "valid" if not errors else "invalid"
        if not errors:
            valid_customer_rows += 1
        db.add(
            models.StagingCustomer(
                company_id=batch.company_id,
                upload_batch_id=batch.id,
                row_number=row_number,
                external_code=external_code,
                full_name=str(payload.get("Nome") or "").strip() or None,
                normalized_name=normalize_text(payload.get("Nome")),
                document_number=document,
                email_billing=parse_email(payload.get("E-mail para cobrança")),
                email_financial=parse_email(payload.get("E-mail do financeiro")),
                phone=str(payload.get("Telefone") or "").strip() or None,
                other_contacts=str(payload.get("Outros contatos") or "").strip() or None,
                raw_payload=safe_json(payload),
                validation_status=validation_status,
                validation_errors=safe_json(errors),
            )
        )

    db.flush()

    customer_stage_rows = (
        db.query(models.StagingCustomer)
        .filter(models.StagingCustomer.upload_batch_id == batch.id, models.StagingCustomer.validation_status == "valid")
        .all()
    )
    customer_index_by_code = {row.external_code: row for row in customer_stage_rows if row.external_code}
    customer_index_by_document = {row.document_number: row for row in customer_stage_rows if row.document_number}

    for row_number, payload in receivable_rows:
        errors = _receivable_validation_errors(payload)
        due_date = parse_date(payload.get("Vencimento"))
        issue_date = parse_date(payload.get("Emissão") or payload.get("Emissao"))
        amount_total = parse_decimal(payload.get("Valor"))
        balance = parse_decimal(payload.get("Saldo")) or Decimal("0.00")
        balance_without_interest = parse_decimal(payload.get("Saldo sem juros")) or balance
        customer_external_code = str(payload.get("Código") or payload.get("Codigo") or "").strip() or None
        customer_document = normalize_digits(payload.get("CNPJ/CPF") or payload.get("Documento do cliente"))
        if customer_external_code is None and customer_document is None:
            errors.append("Cliente sem código ou documento para vinculação segura")

        validation_status = "valid" if not errors else "invalid"
        if not errors:
            valid_receivable_rows += 1

        staging_receivable = models.StagingReceivable(
            company_id=batch.company_id,
            upload_batch_id=batch.id,
            row_number=row_number,
            customer_external_code=customer_external_code,
            customer_name=str(payload.get("Nome") or "").strip() or None,
            normalized_customer_name=normalize_text(payload.get("Nome")),
            customer_document_number=customer_document,
            receivable_number=str(payload.get("Documento") or "").strip() or None,
            nosso_numero=str(payload.get("Nosso Numero") or payload.get("Nosso Número") or "").strip() or None,
            charge_type=str(payload.get("Tipo de cobrança") or payload.get("Tipo") or "").strip() or None,
            issue_date=issue_date,
            due_date=due_date,
            amount_total=amount_total,
            balance_amount=balance,
            balance_without_interest=balance_without_interest,
            status_raw=str(payload.get("Status") or "").strip() or None,
            email_billing=parse_email(payload.get("E-mail para cobrança") or payload.get("E-mail")),
            raw_payload=safe_json(payload),
            validation_status=validation_status,
            validation_errors=safe_json(errors),
        )
        db.add(staging_receivable)
        db.flush()

        if not errors:
            matched_stage_customer = None
            if customer_external_code and customer_external_code in customer_index_by_code:
                matched_stage_customer = customer_index_by_code[customer_external_code]
            elif customer_document and customer_document in customer_index_by_document:
                matched_stage_customer = customer_index_by_document[customer_document]

            matched_customer = None
            if not matched_stage_customer:
                if customer_external_code:
                    matched_customer = (
                        db.query(models.Customer)
                        .filter(
                            models.Customer.company_id == batch.company_id,
                            models.Customer.external_code == customer_external_code,
                            models.Customer.is_active.is_(True),
                        )
                        .first()
                    )
                if not matched_customer and customer_document:
                    matched_customer = (
                        db.query(models.Customer)
                        .filter(
                            models.Customer.company_id == batch.company_id,
                            models.Customer.document_number == customer_document,
                            models.Customer.is_active.is_(True),
                        )
                        .first()
                    )

            if not matched_stage_customer and not matched_customer:
                db.add(
                    models.CustomerLinkPending(
                        company_id=batch.company_id,
                        upload_batch_id=batch.id,
                        staging_receivable_id=staging_receivable.id,
                        status="open",
                    )
                )

    invalid_count = (
        db.query(models.StagingCustomer)
        .filter(models.StagingCustomer.upload_batch_id == batch.id, models.StagingCustomer.validation_status == "invalid")
        .count()
        + db.query(models.StagingReceivable)
        .filter(models.StagingReceivable.upload_batch_id == batch.id, models.StagingReceivable.validation_status == "invalid")
        .count()
    )
    pending_count = (
        db.query(models.CustomerLinkPending)
        .filter(models.CustomerLinkPending.upload_batch_id == batch.id, models.CustomerLinkPending.status == "open")
        .count()
    )
    batch.preview_total_customers = len(customer_rows)
    batch.preview_total_receivables = len(receivable_rows)
    batch.preview_total_pending_links = pending_count
    batch.preview_total_errors = invalid_count
    batch.status = "preview_ready"
    batch.updated_at = datetime.now(timezone.utc)
    if not customer_rows:
        preview_errors.append("Planilha de clientes sem linhas válidas")
    if not receivable_rows:
        preview_errors.append("Planilha de contas a receber sem linhas válidas")
    if preview_errors:
        batch.preview_total_errors += len(preview_errors)


def get_batch_preview(db: Session, batch: models.UploadBatch) -> dict:
    valid_customer_rows = (
        db.query(models.StagingCustomer)
        .filter(models.StagingCustomer.upload_batch_id == batch.id, models.StagingCustomer.validation_status == "valid")
        .count()
    )
    valid_receivable_rows = (
        db.query(models.StagingReceivable)
        .filter(models.StagingReceivable.upload_batch_id == batch.id, models.StagingReceivable.validation_status == "valid")
        .count()
    )
    pending_links = (
        db.query(models.CustomerLinkPending)
        .filter(models.CustomerLinkPending.upload_batch_id == batch.id, models.CustomerLinkPending.status == "open")
        .count()
    )
    errors = []
    if batch.preview_total_errors:
        errors.append("Há linhas inválidas no lote")
    if pending_links:
        errors.append("Há pendências de vinculação manual")
    return {
        "batch_id": batch.id,
        "status": batch.status,
        "total_customer_rows": batch.preview_total_customers,
        "total_receivable_rows": batch.preview_total_receivables,
        "valid_customer_rows": valid_customer_rows,
        "valid_receivable_rows": valid_receivable_rows,
        "pending_links": pending_links,
        "errors": errors,
    }


def _upsert_customer_from_staging(db: Session, company_id: int, row: models.StagingCustomer) -> models.Customer:
    customer = None
    if row.external_code:
        customer = (
            db.query(models.Customer)
            .filter(models.Customer.company_id == company_id, models.Customer.external_code == row.external_code)
            .first()
        )
    if not customer and row.document_number:
        customer = (
            db.query(models.Customer)
            .filter(models.Customer.company_id == company_id, models.Customer.document_number == row.document_number)
            .first()
        )
    if customer:
        customer.full_name = row.full_name or customer.full_name
        customer.normalized_name = row.normalized_name or customer.normalized_name
        customer.document_number = row.document_number or customer.document_number
        customer.email_billing = row.email_billing or customer.email_billing
        customer.email_financial = row.email_financial or customer.email_financial
        customer.phone = row.phone or customer.phone
        customer.other_contacts = row.other_contacts or customer.other_contacts
        customer.is_active = True
        if row.external_code:
            customer.external_code = row.external_code
    else:
        customer = models.Customer(
            company_id=company_id,
            external_code=row.external_code,
            full_name=row.full_name or "Cliente sem nome",
            normalized_name=row.normalized_name or normalize_text(row.full_name),
            document_number=row.document_number,
            email_billing=row.email_billing,
            email_financial=row.email_financial,
            phone=row.phone,
            other_contacts=row.other_contacts,
            is_active=True,
        )
        db.add(customer)
        db.flush()
    return customer


def merge_batch(db: Session, *, batch: models.UploadBatch, approved_by: models.User) -> None:
    pending_open = (
        db.query(models.CustomerLinkPending)
        .filter(models.CustomerLinkPending.upload_batch_id == batch.id, models.CustomerLinkPending.status == "open")
        .count()
    )
    invalid_rows = (
        db.query(models.StagingCustomer)
        .filter(models.StagingCustomer.upload_batch_id == batch.id, models.StagingCustomer.validation_status == "invalid")
        .count()
        + db.query(models.StagingReceivable)
        .filter(models.StagingReceivable.upload_batch_id == batch.id, models.StagingReceivable.validation_status == "invalid")
        .count()
    )
    if pending_open > 0 or invalid_rows > 0:
        raise HTTPException(status_code=400, detail="Lote possui pendências ou linhas inválidas")

    staging_customers = (
        db.query(models.StagingCustomer)
        .filter(models.StagingCustomer.upload_batch_id == batch.id, models.StagingCustomer.validation_status == "valid")
        .all()
    )
    customer_ids_touched: list[int] = []
    for row in staging_customers:
        customer = _upsert_customer_from_staging(db, batch.company_id, row)
        customer_ids_touched.append(customer.id)

    db.flush()

    touched_receivables: list[int] = []
    staging_receivables = (
        db.query(models.StagingReceivable)
        .filter(models.StagingReceivable.upload_batch_id == batch.id, models.StagingReceivable.validation_status == "valid")
        .all()
    )
    for row in staging_receivables:
        customer = None
        if row.customer_external_code:
            customer = (
                db.query(models.Customer)
                .filter(
                    models.Customer.company_id == batch.company_id,
                    models.Customer.external_code == row.customer_external_code,
                    models.Customer.is_active.is_(True),
                )
                .first()
            )
        if not customer and row.customer_document_number:
            customer = (
                db.query(models.Customer)
                .filter(
                    models.Customer.company_id == batch.company_id,
                    models.Customer.document_number == row.customer_document_number,
                    models.Customer.is_active.is_(True),
                )
                .first()
            )
        if not customer:
            raise HTTPException(status_code=400, detail="Existe cobrança válida sem cliente vinculado")

        receivable = None
        if row.nosso_numero:
            receivable = (
                db.query(models.Receivable)
                .filter(models.Receivable.company_id == batch.company_id, models.Receivable.nosso_numero == row.nosso_numero)
                .first()
            )
        if not receivable and row.receivable_number and row.due_date:
            receivable = (
                db.query(models.Receivable)
                .filter(
                    models.Receivable.company_id == batch.company_id,
                    models.Receivable.receivable_number == row.receivable_number,
                    models.Receivable.due_date == row.due_date,
                )
                .first()
            )

        new_status = normalize_receivable_status(row.status_raw, row.due_date, row.balance_amount)
        if receivable:
            old_status = receivable.status
            receivable.customer_id = customer.id
            receivable.upload_batch_id = batch.id
            receivable.receivable_number = row.receivable_number
            receivable.nosso_numero = row.nosso_numero
            receivable.charge_type = row.charge_type
            receivable.issue_date = row.issue_date
            receivable.due_date = row.due_date
            receivable.amount_total = row.amount_total or Decimal("0.00")
            receivable.balance_amount = row.balance_amount or Decimal("0.00")
            receivable.balance_without_interest = row.balance_without_interest or receivable.balance_amount
            receivable.status = new_status
            receivable.customer_name_snapshot = customer.full_name
            receivable.billing_email_snapshot = customer.email_billing
            receivable.document_snapshot = customer.document_number
            receivable.is_active = True
            if old_status != new_status:
                db.add(
                    models.ReceivableHistory(
                        company_id=batch.company_id,
                        receivable_id=receivable.id,
                        event_type="status_change",
                        old_status=old_status,
                        new_status=new_status,
                        note="Atualizado pelo merge do lote",
                        created_by_user_id=approved_by.id,
                    )
                )
        else:
            receivable = models.Receivable(
                company_id=batch.company_id,
                customer_id=customer.id,
                upload_batch_id=batch.id,
                receivable_number=row.receivable_number,
                nosso_numero=row.nosso_numero,
                charge_type=row.charge_type,
                issue_date=row.issue_date,
                due_date=row.due_date,
                amount_total=row.amount_total or Decimal("0.00"),
                balance_amount=row.balance_amount or Decimal("0.00"),
                balance_without_interest=row.balance_without_interest or row.balance_amount or Decimal("0.00"),
                status=new_status,
                customer_name_snapshot=customer.full_name,
                billing_email_snapshot=customer.email_billing,
                document_snapshot=customer.document_number,
                is_active=True,
            )
            db.add(receivable)
            db.flush()
            db.add(
                models.ReceivableHistory(
                    company_id=batch.company_id,
                    receivable_id=receivable.id,
                    event_type="created",
                    new_status=new_status,
                    note="Criado pelo merge do lote",
                    created_by_user_id=approved_by.id,
                )
            )
        touched_receivables.append(receivable.id)

    cutoff = datetime.now(timezone.utc) - timedelta(days=settings.old_data_inactivation_days)
    if touched_receivables:
        db.query(models.Receivable).filter(
            models.Receivable.company_id == batch.company_id,
            models.Receivable.is_active.is_(True),
            ~models.Receivable.id.in_(touched_receivables),
            models.Receivable.updated_at < cutoff,
            models.Receivable.status.notin_(["pago", "cancelado"]),
        ).update({models.Receivable.is_active: False}, synchronize_session=False)

    batch.status = "merged"
    batch.approved_at = datetime.now(timezone.utc)
    batch.approved_by_user_id = approved_by.id
    batch.updated_at = datetime.now(timezone.utc)

    db.add(
        models.AuditLog(
            company_id=batch.company_id,
            user_id=approved_by.id,
            entity_type="upload_batch",
            entity_id=str(batch.id),
            action="merge_approved",
            details=safe_json({"batch_reference": batch.batch_reference}),
        )
    )


def create_customer_from_pending(
    db: Session,
    *,
    pending: models.CustomerLinkPending,
    payload: dict,
    user: models.User,
) -> models.Customer:
    customer = models.Customer(
        company_id=user.company_id,
        full_name=payload["full_name"].strip(),
        normalized_name=normalize_text(payload["full_name"]),
        document_number=normalize_digits(payload.get("document_number")),
        email_billing=(payload.get("email_billing") or None),
        email_financial=(payload.get("email_financial") or None),
        phone=(payload.get("phone") or None),
        other_contacts=(payload.get("other_contacts") or None),
        is_active=True,
    )
    db.add(customer)
    db.flush()
    resolve_pending_with_customer(db, pending=pending, customer=customer, user=user, note="Perfil criado manualmente")
    return customer


def resolve_pending_with_customer(
    db: Session,
    *,
    pending: models.CustomerLinkPending,
    customer: models.Customer,
    user: models.User,
    note: str,
) -> None:
    staging = db.query(models.StagingReceivable).filter(models.StagingReceivable.id == pending.staging_receivable_id).first()
    if not staging:
        raise HTTPException(status_code=404, detail="Registro de staging não encontrado")
    if customer.company_id != user.company_id:
        raise HTTPException(status_code=403, detail="Cliente fora da empresa do usuário")
    if staging.customer_external_code is None and customer.external_code:
        staging.customer_external_code = customer.external_code
    if staging.customer_document_number is None and customer.document_number:
        staging.customer_document_number = customer.document_number
    pending.status = "resolved"
    pending.suggested_customer_id = customer.id
    pending.resolution_note = note
    pending.resolved_by_user_id = user.id
    pending.resolved_at = datetime.now(timezone.utc)
    db.add(
        models.AuditLog(
            company_id=user.company_id,
            user_id=user.id,
            entity_type="customer_link_pending",
            entity_id=str(pending.id),
            action="resolved",
            details=safe_json({"customer_id": customer.id, "note": note}),
        )
    )


def cleanup_old_staging(db: Session, company_id: int) -> int:
    threshold = datetime.now(timezone.utc) - timedelta(days=settings.staging_retention_days)
    old_batches = db.query(models.UploadBatch).filter(
        models.UploadBatch.company_id == company_id,
        models.UploadBatch.created_at < threshold,
        models.UploadBatch.status.in_(["preview_ready", "merged", "uploaded"]),
    )
    count = 0
    for batch in old_batches.all():
        _clear_batch_staging(db, batch.id)
        count += 1
    return count